import openpyxl
from models import Wrestler, Team

def load_all_americans_from_excel(file_path):
    """Load All American, finalist, and medal wrestlers from an Excel file."""
    all_americans = set()
    finalists = set()
    medals = {}
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for cell in sheet.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True):
        for name in cell:
            if name:
                all_americans.add(name)

    for cell in sheet.iter_cols(min_col=2, max_col=2, min_row=2, values_only=True):
        for name in cell:
            if name:
                finalists.add(name)

    for col in sheet.iter_cols(min_col=3, values_only=True):
        medal = col[0]
        for name in col[1:]:
            if name:
                medals[name] = medal

    return all_americans, finalists, medals

def load_wrestlers_from_excel(file_path):
    """Load wrestlers from an Excel file."""
    wrestlers = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for col in sheet.iter_cols(min_row=0, min_col=1, values_only=True):
        weight_class = col[0]
        for cell in col[1:]:
            if cell:
                seed, name = cell.split(') ')
                seed = int(seed.strip('('))
                last_name, first_name = name.split(', ')
                full_name = f"{first_name} {last_name}"
                wrestler = Wrestler(full_name, seed, weight_class)
                wrestlers.append(wrestler)

    return wrestlers

def load_teams_from_excel(file_path):
    """Load teams from an Excel file."""
    teams = []
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for col in sheet.iter_cols(min_row=0, min_col=2, values_only=True):
        team_name = col[0]
        tie_breaker_team = col[1]
        tie_breaker_score = col[2]
        team = Team(team_name, 0, tie_breaker_team, tie_breaker_score)
        teams.append(team)

    return teams

def load_teams_wrestlers_from_excel(file_path):
    """Load wrestlers for each team from an Excel file."""
    teams_wrestlers = {}
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for col in sheet.iter_cols(min_row=1, min_col=1, values_only=True):
        team_name = col[0]
        wrestlers = []
        for cell in col[1:]:
            if cell:
                seed, name = cell.split(') ')
                seed = int(seed.strip('('))
                wrestler = Wrestler(name, seed, None)
                wrestlers.append(wrestler)
        teams_wrestlers[team_name] = wrestlers

    return teams_wrestlers

# Example usage
if __name__ == "__main__":
    wrestlers_file_path = "wrestlers.xlsx"
    wrestlers = load_wrestlers_from_excel(wrestlers_file_path)
    for wrestler in wrestlers:
        print(f"Name: {wrestler.name}, Seed: {wrestler.seed}, Weight Class: {wrestler.weight_class}")

    teams_file_path = "teams.xlsx"
    teams = load_teams_from_excel(teams_file_path)
    for team in teams:
        print(f"Team: {team.name}, Tie Breaker Team: {team.tie_breaker_team}, Tie Breaker Score: {team.tie_breaker_score}")